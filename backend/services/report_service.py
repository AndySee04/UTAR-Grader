from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime
import math
from xml.sax.saxutils import escape

PAGE_W, _PAGE_H = A4

# Same bands as frontend ExamResults.jsx (lexical confidence 0–1).
_CONF_COLORS = {"high": "#047857", "medium": "#b45309", "low": "#be123c"}


def _lexical_confidence_level(confidence: Any) -> Optional[Dict[str, Any]]:
    """High ≥0.85, Medium 0.65–0.84, Low <0.65."""
    if confidence is None:
        return None
    try:
        x = float(confidence)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    pct = int(round(x * 100))
    if x >= 0.85:
        return {"label": "High", "pct": pct, "band": "high"}
    if x >= 0.65:
        return {"label": "Medium", "pct": pct, "band": "medium"}
    return {"label": "Low", "pct": pct, "band": "low"}


def _pdf_paragraph(text: str, style: ParagraphStyle, *, empty_placeholder: str = "—") -> Paragraph:
    """Build a ReportLab Paragraph from plain text (escape HTML, preserve line breaks)."""
    s = (text or "").strip()
    if not s:
        return Paragraph(f"<i>{escape(empty_placeholder)}</i>", style)
    return Paragraph(escape(s).replace("\n", "<br/>"), style)


class ReportService:
    """Service for generating Excel and PDF reports."""

    def generate_excel_summary(
        self,
        exam_name: str,
        students: List[Dict[str, Any]],
        questions: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate Excel summary report for all students.

        Args:
            exam_name: Name of the exam
            students: List of student grade summaries
            questions: List of questions with max marks

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Grade Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Grade Report: {exam_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        headers = ["Student Name"] + [q.get("question_number", f"Q{i+1}") for i, q in enumerate(questions)] + ["Total", "Percentage"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        ws.cell(row=5, column=1, value="Max Marks").font = Font(italic=True)
        total_max = 0
        for col, q in enumerate(questions, 2):
            max_marks = q.get("max_marks", 0)
            ws.cell(row=5, column=col, value=max_marks).border = border
            total_max += max_marks
        ws.cell(row=5, column=len(questions) + 2, value=total_max).border = border
        ws.cell(row=5, column=len(questions) + 3, value="100%").border = border

        for row, student in enumerate(students, 6):
            ws.cell(row=row, column=1, value=student.get("student_name", "Unknown")).border = border

            grades = student.get("grades", [])
            grade_by_q = {g.get("question_number"): g for g in grades}

            for col, q in enumerate(questions, 2):
                q_num = q.get("question_number")
                grade = grade_by_q.get(q_num, {})
                score = grade.get("score", 0)
                ws.cell(row=row, column=col, value=score).border = border

            ws.cell(row=row, column=len(questions) + 2, value=student.get("total_score", 0)).border = border
            ws.cell(row=row, column=len(questions) + 3, value=f"{student.get('percentage', 0):.1f}%").border = border

        ws.column_dimensions['A'].width = 20
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 10

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_student_pdf(
        self,
        exam_name: str,
        student_name: str,
        grades: List[Dict[str, Any]],
        total_score: float,
        total_max: float,
        percentage: float
    ) -> bytes:
        """
        Generate PDF report for a single student: per question, question-only line,
        rule, then a student-answer row (label + marks/confidence) and boxes for
        answer text and AI feedback.
        """
        LM = 0.62 * inch
        RM = 0.62 * inch
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=LM,
            rightMargin=RM,
            topMargin=0.55 * inch,
            bottomMargin=0.55 * inch,
            title="Grade Report",
        )
        content_w = PAGE_W - LM - RM

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "PdfTitle",
            parent=styles["Title"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=8,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        meta_style = ParagraphStyle(
            "PdfMeta",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )
        summary_cell_style = ParagraphStyle(
            "SumCell",
            parent=styles["Normal"],
            fontSize=14,
            leading=17,
            textColor=colors.HexColor("#0f172a"),
            alignment=TA_CENTER,
            spaceBefore=0,
            spaceAfter=0,
        )
        q_row_left_style = ParagraphStyle(
            "QRowLeft",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#0f172a"),
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
        )
        q_row_right_style = ParagraphStyle(
            "QRowRight",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#0f172a"),
            alignment=TA_RIGHT,
            spaceBefore=0,
            spaceAfter=0,
        )
        section_label = ParagraphStyle(
            "SecLabel",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
            alignment=TA_LEFT,
            spaceBefore=8,
            spaceAfter=4,
            fontName="Helvetica-Bold",
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=9,
            leading=13,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#1e293b"),
        )
        def _sort_key(g: Dict[str, Any]) -> str:
            return str(g.get("question_number") or "")

        grades_sorted = sorted(grades, key=_sort_key)

        elements: List[Any] = []

        elements.append(Paragraph(f"Grade Report: {escape(exam_name)}", title_style))
        elements.append(Paragraph(f"<b>Student:</b> {escape(student_name or 'Unknown')}", meta_style))
        elements.append(
            Paragraph(
                f"<b>Generated:</b> {escape(datetime.now().strftime('%Y-%m-%d %H:%M'))}",
                meta_style,
            )
        )
        elements.append(Spacer(1, 14))

        summary_inner = Table(
            [
                [
                    Paragraph(
                        f'<font size="9" color="#64748b">Total score</font><br/>'
                        f'<b><font size="14" color="#0f172a">{escape(f"{total_score:.1f} / {total_max:.1f}")}</font></b>',
                        summary_cell_style,
                    ),
                    Paragraph(
                        f'<font size="9" color="#64748b">Percentage</font><br/>'
                        f'<b><font size="14" color="#0f172a">{escape(f"{percentage:.1f}%")}</font></b>',
                        summary_cell_style,
                    ),
                ],
            ],
            colWidths=[content_w * 0.5, content_w * 0.5],
        )
        summary_inner.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f5f9")),
                    ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#cbd5e1")),
                    ("LINEBEFORE", (1, 0), (1, 0), 1, colors.HexColor("#cbd5e1")),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(summary_inner)
        elements.append(Spacer(1, 22))

        sa_heading_left_style = ParagraphStyle(
            "SaHeadLeft",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
            fontName="Helvetica-Bold",
        )

        for grade in grades_sorted:
            qn = str(grade.get("question_number") or "—")
            qt = (grade.get("question_text") or "").strip()
            score = float(grade.get("score") or 0)
            mx = float(grade.get("max_marks") or 0)
            c = grade.get("confidence")
            score_s = f"{score:.1f} / {mx:.1f}"

            if qt:
                left_html = f"<b>{escape(qn)}</b> {escape(qt)}"
            else:
                left_html = f"<b>{escape(qn)}</b>"
            question_only = Table(
                [[Paragraph(left_html, q_row_left_style)]],
                colWidths=[content_w],
            )
            question_only.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(question_only)

            hr = Table([[""]], colWidths=[content_w])
            hr.setStyle(
                TableStyle(
                    [
                        ("LINEABOVE", (0, 0), (-1, -1), 0.55, colors.HexColor("#1e293b")),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(hr)

            lvl = _lexical_confidence_level(c)
            if lvl:
                col_hex = _CONF_COLORS[lvl["band"]]
                marks_para = Paragraph(
                    f'<font color="{col_hex}"><b>{escape(lvl["label"])}</b></font>'
                    f' <font size="8" color="#64748b">({lvl["pct"]}%)</font>'
                    f' &nbsp;&nbsp;<b>{escape(score_s)}</b>',
                    q_row_right_style,
                )
            else:
                marks_para = Paragraph(f"<b>{escape(score_s)}</b>", q_row_right_style)

            sa_title_row = Table(
                [
                    [
                        Paragraph("<b>Student answer</b>", sa_heading_left_style),
                        marks_para,
                    ]
                ],
                colWidths=[content_w * 0.52, content_w * 0.48],
            )
            sa_title_row.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(sa_title_row)
            sa_para = _pdf_paragraph(
                grade.get("student_answer") or "",
                body_style,
                empty_placeholder="No answer text on file.",
            )
            sa_box = Table([[sa_para]], colWidths=[content_w])
            sa_box.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#e2e8f0")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 12),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                        ("TOPPADDING", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(sa_box)

            elements.append(Paragraph("AI feedback", section_label))
            fb_para = _pdf_paragraph(
                grade.get("feedback") or "",
                body_style,
                empty_placeholder="No feedback.",
            )
            fb_box = Table([[fb_para]], colWidths=[content_w])
            fb_box.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef2ff")),
                        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#a5b4fc")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 12),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                        ("TOPPADDING", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(fb_box)

            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


report_service = ReportService()
